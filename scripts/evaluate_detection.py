#!/usr/bin/env python3
#
# Copyright (c) 2026 Direction générale de la prévention des risques (DGPR).
#
# This file is part of OCAPI.
# See https://github.com/mte-dgpr/ocapi for further info.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
"""Evaluate LLM operation detection against ground-truth annotations.

Usage (from repo root):
  python scripts/evaluate_detection.py --model openai_gpt5mini
  python scripts/evaluate_detection.py --model mistral_medium --aiot 0003013459
  python scripts/evaluate_detection.py --model mistral_medium --save-ops --save-score
  python scripts/evaluate_detection.py --score-only eval/2026-06-05_10-00_eval_mistral_medium
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

from ocapi.exceptions import InputOutputError  # noqa: E402
from ocapi.llm_utils import (  # noqa: E402
    TokenUsage,
    config_model_llm,
    get_accumulated_usage,
    reset_accumulated_usage,
)
from ocapi.step_detection import step_detection as step_detection_module  # noqa: E402
from ocapi.step_detection.step_detection import _OPERATION_ID_COUNTER, step_detection  # noqa: E402
from ocapi.step_resolution.build_op_graph import build_graph  # noqa: E402
from ocapi.step_tagging import step_tagging  # noqa: E402
from ocapi.step_tagging.operations_filtering import filter_redundant_operations  # noqa: E402
from ocapi.types import Operation, is_low_severity_op, numbering_fragment_for_sort  # noqa: E402
from ocapi.utils.io_utils import (  # noqa: E402
    load_arrete_files,
    load_document_contexts,
    load_operations,
    save_operations,
)
from ocapi.utils.logging_utils import get_logger, initialize_root_logger  # noqa: E402
from ocapi.utils.tagging_io import extract_operations_from_tagged_soup  # noqa: E402
from ocapi.utils.utils import make_id  # noqa: E402

_LOGGER = get_logger(__name__)

_GROUND_TRUTH_DIR = _PROJECT_ROOT / "snapshots" / "ground-truth"
_ARRETES_HTML_DIR = _PROJECT_ROOT / "snapshots" / "arretes_html"

_VALID_OP_TYPES = {"ADD", "REPLACE", "REMOVE"}

# Cost per 1M tokens (USD): {model_id: (input_cost, output_cost)}
_COST_PER_1M_TOKENS: dict[str, tuple[float, float]] = {
    "gpt-5.4": (2.50, 15.00),
    "gpt-5.4-mini": (0.75, 4.50),
    "gpt-5.4-nano": (0.20, 1.25),
    "gpt-5": (1.25, 10.00),
    "gpt-5-mini": (0.25, 2.00),
    "gpt-5-nano": (0.05, 0.40),
    "gpt-4o": (2.50, 10.00),
    "mte-api-piag-mistral-medium-latest": (2.00, 6.00),
    "mistral-large-2512": (0.50, 1.50),
    "mistral-medium-3-5": (1.50, 7.50),
    "mistral-medium-2508": (0.40, 2.00),
    "mistral-small-2603": (0.15, 0.6),
    "claude-opus-4-8": (5.00, 25.00),
    "claude-sonnet-4-6": (3.00, 15.00),
    "claude-haiku-4-5": (1.00, 5.00),
    "gemini-3.5-flash": (1.50, 9.00),
    "gemini-3.1-pro-preview": (4.00, 18.00),
    "gemini-3.1-flash-lite": (0.25, 1.50),
    "gemini-2.5-pro": (2.50, 10.00),
    "deepseek-v4-pro": (0.44, 0.87),
    "deepseek-v4-flash": (0.14, 0.28),
}


def _compute_cost(model_id: str, usage: TokenUsage) -> float:
    """Compute cost in USD from token usage."""
    rates = _COST_PER_1M_TOKENS.get(model_id, (0.0, 0.0))
    return (usage.prompt_tokens * rates[0] + usage.completion_tokens * rates[1]) / 1_000_000


# (source_arrete, source_article, target_arrete, target_article, operation_type)
OperationKey = tuple[str, str, str, str, str]


def _operation_key_from_dict(op: dict) -> OperationKey:
    """Convert operation dict from JSON to OperationKey tuple, applying numbering fragment
    extraction."""
    return (
        op["source_id"]["arrete_id"],
        numbering_fragment_for_sort(op["source_id"]["article_id"]),
        op["target_id"]["arrete_id"],
        numbering_fragment_for_sort(op["target_id"]["article_id"]),
        op["operation_type"],
    )


def _operation_key(op: Operation | dict) -> OperationKey:
    """Convert an Operation (or dict) to an OperationKey tuple, applying numbering fragment
    extraction."""
    if isinstance(op, dict):
        return _operation_key_from_dict(op)

    operation_type = op.operation_type
    if hasattr(operation_type, "value"):
        operation_type_value = operation_type.value
    else:
        operation_type_value = str(operation_type)

    return (
        op.source_id.arrete_id,
        numbering_fragment_for_sort(op.source_id.article_id),
        op.target_id.arrete_id,
        numbering_fragment_for_sort(op.target_id.article_id),
        operation_type_value,
    )


def load_ground_truth(aiot: str) -> list[OperationKey]:
    """Load ground-truth operations for an AIOT, filtering out AUTRE."""
    gt_path = _GROUND_TRUTH_DIR / aiot / "operations.json"
    raw = json.loads(gt_path.read_text(encoding="utf-8"))
    return [_operation_key_from_dict(op) for op in raw if op["operation_type"] in _VALID_OP_TYPES]


def compare_operations(
    detected: list[OperationKey],
    ground_truth: list[OperationKey],
) -> tuple[int, int, int]:
    """Compare detected operations with ground-truth using multiset matching.

    Returns (true_positives, false_positives, false_negatives).
    """
    detected_counts = Counter(detected)
    gt_counts = Counter(ground_truth)

    tp = sum(min(detected_counts[k], gt_counts[k]) for k in detected_counts if k in gt_counts)
    fp = sum(detected_counts.values()) - tp
    fn = sum(gt_counts.values()) - tp
    return tp, fp, fn


@dataclass
class Scores:
    precision: float
    recall: float
    f1: float


def compute_scores(tp: int, fp: int, fn: int) -> Scores:
    """Compute precision, recall, and F1-score from raw counts."""
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    return Scores(precision=precision, recall=recall, f1=f1)


def run_detection_for_aiot(
    aiot: str,
    model_key: str,
    *,
    enable_tagging: bool = True,
    enable_tagging_ops: bool = False,
) -> tuple[list[Operation], list[Operation]]:
    """Run optional tagging + detection on all arrêtés of an AIOT.

    When ``enable_tagging`` is true, the script first runs :func:`step_tagging`
    on each loaded document context. When ``enable_tagging_ops`` is also true,
    it extracts the regex-tagged operations, then runs LLM detection and merges
    both sources with the same policy as the main pipeline. This makes it
    possible to evaluate the detection-only baseline or the combined tagging +
    detection flow on the same AIOT.

    Returns
    -------
    tuple[list[Operation], list[Operation]]
        A pair ``(raw_ops, validated_ops)`` where ``raw_ops`` is the full list
        of operations returned by the LLM and ``validated_ops`` is the subset
        after running :func:`build_graph` (which assigns context-dependent error
        codes such as ``MISSING_ARRETE`` or ``LESS_IMPORTANT``) and filtering
        out operations that carry only LOW-severity error codes.
    """
    arretes_dir = _ARRETES_HTML_DIR / aiot
    pairs = load_document_contexts(arretes_dir, aiot)
    arrete_files = [arrete_file for arrete_file, _ in pairs]
    document_contexts = [document_context for _, document_context in pairs]
    if not arrete_files:
        _LOGGER.warning(f"No arrêtés loaded for {aiot}")
        return [], []

    step_detection_module.LLM_CFG = config_model_llm(model_key)
    # Keep IDs deterministic per AIOT, same behavior as run_pipeline.
    _OPERATION_ID_COUNTER.value = 0

    tagged_ops: list[Operation] = []
    if enable_tagging:
        for arrete_file, document_context in zip(arrete_files, document_contexts):
            step_tagging(document_context)
            arrete_file.soup = document_context.soup
    if enable_tagging_ops:
        for arrete_file in arrete_files:
            tagged_ops.extend(
                extract_operations_from_tagged_soup(
                    arrete_file.soup,
                    arrete_file.id,
                    next_operation_id=lambda: make_id(_OPERATION_ID_COUNTER),
                )
            )

    start_date = arrete_files[0].id
    detected_ops: list[Operation] = []
    for arrete_file in arrete_files:
        if arrete_file.id <= start_date:
            continue
        file_ops = step_detection(arrete_file)
        detected_ops.extend(file_ops)
        _LOGGER.info(f"  {arrete_file.id}: {len(file_ops)} operations detected")

    if enable_tagging_ops:
        # Same merge policy as pipeline: keep the most precise sub-target when
        # tagging and LLM disagree on granularity; renumber collisions on kept ops.
        detected_ops = filter_redundant_operations(
            reference_ops=tagged_ops,
            candidate_ops=detected_ops,
            context_id=f"evaluate_detection:{aiot}",
            next_operation_id=lambda: make_id(_OPERATION_ID_COUNTER),
        )

    _LOGGER.info(f"Running build_graph on {len(detected_ops)} operations to assign context errors…")
    _, _, _, updated_ops = build_graph(detected_ops, arrete_files)
    validated_ops = [op for op in updated_ops if not is_low_severity_op(op)]
    _LOGGER.info(
        f"  {len(detected_ops)} raw → {len(updated_ops)} updated → "
        f"{len(validated_ops)} after LOW-severity filter"
    )
    return detected_ops, validated_ops


def _available_aiots() -> list[str]:
    if not _GROUND_TRUTH_DIR.exists():
        return []
    return sorted(d.name for d in _GROUND_TRUTH_DIR.iterdir() if d.is_dir())


def _validated_operation_keys(
    detected_ops: list[Operation],
    arrete_files: list,
) -> tuple[list[OperationKey], list[Operation]]:
    """Return validated operation keys and the filtered operations used for scoring."""
    _, _, _, updated_ops = build_graph(detected_ops, arrete_files)
    validated_ops = [op for op in updated_ops if not is_low_severity_op(op)]
    return [_operation_key(op) for op in validated_ops], validated_ops


@dataclass
class AiotResult:
    aiot: str
    gt_count: int
    detected_count: int
    validated_count: int
    tp: int
    fp: int
    fn: int
    scores: Scores
    elapsed_seconds: float = 0.0
    usage: TokenUsage = field(default_factory=TokenUsage)
    cost_usd: float = 0.0

    def __post_init__(self) -> None:
        """Snapshot token usage to keep per-AIOT metrics stable in exported reports."""
        self.usage = TokenUsage(
            prompt_tokens=self.usage.prompt_tokens,
            completion_tokens=self.usage.completion_tokens,
        )


def _write_xlsx(
    results: list[AiotResult],
    overall: Scores,
    total_tp: int,
    total_fp: int,
    total_fn: int,
    model_key: str,
    output_path: Path,
) -> None:

    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pct_fmt = "0.0%"

    headers = [
        "AIOT",
        "Ground-truth",
        "Détectées",
        "Validées",
        "TP",
        "FP",
        "FN",
        "Precision",
        "Recall",
        "F1",
        "Temps (s)",
        "Tokens in",
        "Tokens out",
        "Coût ($)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.aiot)
        ws.cell(row=row_idx, column=2, value=r.gt_count)
        ws.cell(row=row_idx, column=3, value=r.detected_count)
        ws.cell(row=row_idx, column=4, value=r.validated_count)
        ws.cell(row=row_idx, column=5, value=r.tp)
        ws.cell(row=row_idx, column=6, value=r.fp)
        ws.cell(row=row_idx, column=7, value=r.fn)
        ws.cell(row=row_idx, column=8, value=r.scores.precision).number_format = pct_fmt
        ws.cell(row=row_idx, column=9, value=r.scores.recall).number_format = pct_fmt
        ws.cell(row=row_idx, column=10, value=r.scores.f1).number_format = pct_fmt
        ws.cell(row=row_idx, column=11, value=round(r.elapsed_seconds, 1))
        ws.cell(row=row_idx, column=12, value=r.usage.prompt_tokens)
        ws.cell(row=row_idx, column=13, value=r.usage.completion_tokens)
        ws.cell(row=row_idx, column=14, value=round(r.cost_usd, 4)).number_format = "0.0000"

    total_row = len(results) + 2
    total_font = Font(bold=True)
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=5, value=total_tp).font = total_font
    ws.cell(row=total_row, column=6, value=total_fp).font = total_font
    ws.cell(row=total_row, column=7, value=total_fn).font = total_font
    ws.cell(row=total_row, column=8, value=overall.precision).number_format = pct_fmt
    ws.cell(row=total_row, column=8).font = total_font
    ws.cell(row=total_row, column=9, value=overall.recall).number_format = pct_fmt
    ws.cell(row=total_row, column=9).font = total_font
    ws.cell(row=total_row, column=10, value=overall.f1).number_format = pct_fmt
    ws.cell(row=total_row, column=10).font = total_font
    total_time = sum(r.elapsed_seconds for r in results)
    total_cost = sum(r.cost_usd for r in results)
    total_in = sum(r.usage.prompt_tokens for r in results)
    total_out = sum(r.usage.completion_tokens for r in results)
    ws.cell(row=total_row, column=11, value=round(total_time, 1)).font = total_font
    ws.cell(row=total_row, column=12, value=total_in).font = total_font
    ws.cell(row=total_row, column=13, value=total_out).font = total_font
    c = ws.cell(row=total_row, column=14, value=round(total_cost, 4))
    c.number_format = "0.0000"
    c.font = total_font

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    wb.save(output_path)


def _log_and_accumulate(
    r: AiotResult,
    total_tp: int,
    total_fp: int,
    total_fn: int,
) -> tuple[int, int, int]:
    _LOGGER.info("\n--- %s ---", r.aiot)
    _LOGGER.info(
        "  Ground-truth: %d  |  Detected: %d  |  Validated: %d",
        r.gt_count,
        r.detected_count,
        r.validated_count,
    )
    _LOGGER.info("  TP=%d  FP=%d  FN=%d", r.tp, r.fp, r.fn)
    _LOGGER.info("  Precision: %.3f", r.scores.precision)
    _LOGGER.info("  Recall:    %.3f", r.scores.recall)
    _LOGGER.info("  F1:        %.3f", r.scores.f1)
    return total_tp + r.tp, total_fp + r.fp, total_fn + r.fn


def _make_eval_subdir(model_key: str, ts: datetime) -> Path:
    """Return (and create) eval/<yyyy-mm-dd_hh-mm>_eval_<model> under the project root."""
    ts_str = ts.strftime("%Y-%m-%d_%H-%M")
    safe_model = model_key.replace("/", "-").replace(":", "-")
    folder_name = f"{ts_str}_eval_{safe_model}"
    eval_subdir = _PROJECT_ROOT / "eval" / folder_name
    eval_subdir.mkdir(parents=True, exist_ok=True)
    return eval_subdir


def _run_score_mode(args: argparse.Namespace, ops_dir: Path) -> int:
    """Recompute metrics from existing ops.json files without any LLM call."""
    # Determine which AIOTs to score: either from --aiot or all subdirs with ops.json
    if args.aiot:
        missing = [aiot for aiot in args.aiot if not (ops_dir / aiot / "operations.json").exists()]
        if missing:
            for aiot in missing:
                _LOGGER.error("operations.json not found for AIOT %r in %s", aiot, ops_dir)
            return 1
        aiots = list(args.aiot)
    else:
        if not ops_dir.exists():
            _LOGGER.error("--score-only dir not found: %s", ops_dir)
            return 1
        if not ops_dir.is_dir():
            _LOGGER.error("--score-only is not a directory: %s", ops_dir)
            return 1
        aiots = sorted(
            d.name for d in ops_dir.iterdir() if d.is_dir() and (d / "operations.json").exists()
        )

    if not aiots:
        _LOGGER.error("No operations.json found in %s", ops_dir)
        return 1

    _LOGGER.info(f"Score-only mode — ops-dir: {ops_dir}")
    _LOGGER.info(f"AIOTs: {', '.join(aiots)}")

    total_tp, total_fp, total_fn = 0, 0, 0
    results: list[AiotResult] = []

    for aiot in aiots:
        _LOGGER.info(f"\n{'=' * 50}")
        _LOGGER.info(f"AIOT: {aiot}")
        _LOGGER.info(f"{'=' * 50}")

        gt_keys = load_ground_truth(aiot)
        detected_ops = load_operations(ops_dir / aiot)
        _LOGGER.info(f"Ground-truth: {len(gt_keys)}  |  Loaded ops: {len(detected_ops)}")

        try:
            arrete_files = load_arrete_files(_ARRETES_HTML_DIR / aiot, aiot)
        except InputOutputError:
            arrete_files = []
        if arrete_files:
            validated_keys, validated_ops = _validated_operation_keys(detected_ops, arrete_files)
            _LOGGER.info(
                f"  {len(detected_ops)} loaded → {len(validated_ops)} after LOW-severity filter"
            )
        else:
            _LOGGER.warning(f"No arrêtés found for {aiot} — skipping LOW-severity filtering")
            validated_keys = [_operation_key(op) for op in detected_ops]
            validated_ops = detected_ops

        _LOGGER.info(f"Validated ops for scoring: {len(validated_keys)}")

        tp, fp, fn = compare_operations(validated_keys, gt_keys)
        scores = compute_scores(tp, fp, fn)

        r = AiotResult(
            aiot,
            len(gt_keys),
            len(detected_ops),
            len(validated_ops),
            tp,
            fp,
            fn,
            scores,
        )
        results.append(r)
        total_tp, total_fp, total_fn = _log_and_accumulate(r, total_tp, total_fp, total_fn)

    overall = compute_scores(total_tp, total_fp, total_fn)
    _LOGGER.info("\n%s", "=" * 50)
    _LOGGER.info("OVERALL (score-only mode — %s)", args.model)
    _LOGGER.info("%s", "=" * 50)
    _LOGGER.info("  TP=%d  FP=%d  FN=%d", total_tp, total_fp, total_fn)
    _LOGGER.info("  Precision: %.3f", overall.precision)
    _LOGGER.info("  Recall:    %.3f", overall.recall)
    _LOGGER.info("  F1:        %.3f", overall.f1)

    if args.save_score:
        ts = datetime.now(tz=timezone.utc)
        eval_subdir = _make_eval_subdir(args.model, ts)
        xlsx_path = eval_subdir / "scores.xlsx"
        _write_xlsx(results, overall, total_tp, total_fp, total_fn, args.model, xlsx_path)
        _LOGGER.info("\nResults exported to %s", xlsx_path)

    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Evaluate LLM operation detection against ground-truth.",
    )
    parser.add_argument(
        "--model",
        required=True,
        help="Model key from config/llm_models.json (e.g. openai_gpt5mini, mistral_medium)",
    )
    parser.add_argument(
        "--aiot",
        nargs="*",
        help="AIOT(s) to evaluate (default: all available ground-truth AIOTs)",
    )
    parser.add_argument(
        "--save-ops",
        action="store_true",
        help=("Save detected operations as JSON in eval/<date>_eval_<model>/ after each LLM run."),
    )
    parser.add_argument(
        "--score-only",
        type=Path,
        metavar="EVAL_DIR",
        default=None,
        help=(
            "Score-only mode: recompute metrics from existing ops.json files in EVAL_DIR "
            "(no LLM call)."
        ),
    )
    parser.add_argument(
        "--save-score",
        action="store_true",
        help="Export results to XLSX in eval/<date>_eval_<model>/scores.xlsx.",
    )
    parser.add_argument(
        "--no-tagging",
        action="store_true",
        help="Disable tagging before LLM detection (enabled by default).",
    )
    parser.add_argument(
        "--tagging-ops",
        action="store_true",
        help=(
            "Also extract regex-tagged operations and merge them with candidate ops "
            "(disabled by default)."
        ),
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Enable verbose logging",
    )
    args = parser.parse_args(argv)

    initialize_root_logger(level="DEBUG" if args.verbose else "INFO", console_output=True)

    # --- score-only mode ---
    if args.score_only is not None:
        return _run_score_mode(args, args.score_only)

    # --- detection mode ---

    aiots = args.aiot or _available_aiots()
    if not aiots:
        _LOGGER.error("No AIOT with ground-truth found.")
        return 1

    model_key = args.model
    _LOGGER.info(f"Model: {model_key}")
    _LOGGER.info(f"AIOTs: {', '.join(aiots)}")

    # Create eval subfolder once if any save option is requested
    eval_subdir: Path | None = None
    if args.save_ops or args.save_score:
        eval_subdir = _make_eval_subdir(model_key, datetime.now(tz=timezone.utc))
        _LOGGER.info(f"Eval output directory: {eval_subdir}")

    total_tp, total_fp, total_fn = 0, 0, 0
    results: list[AiotResult] = []

    model_cfg = config_model_llm(model_key)
    model_id = model_cfg.model_name

    for aiot in aiots:
        _LOGGER.info(f"\n{'=' * 50}")
        _LOGGER.info(f"AIOT: {aiot}")
        _LOGGER.info(f"{'=' * 50}")

        gt_keys = load_ground_truth(aiot)
        _LOGGER.info(f"Ground-truth: {len(gt_keys)} operations")

        reset_accumulated_usage()
        t0 = time.monotonic()
        detected_ops, validated_ops = run_detection_for_aiot(
            aiot,
            model_key,
            enable_tagging=not args.no_tagging,
            enable_tagging_ops=args.tagging_ops,
        )
        elapsed = time.monotonic() - t0
        usage = get_accumulated_usage()
        cost = _compute_cost(model_id, usage)

        if args.save_ops and eval_subdir is not None:
            out_dir = eval_subdir / aiot
            save_operations(detected_ops, out_dir)
            _LOGGER.info(f"Ops saved to {out_dir / 'operations.json'}")

        validated_keys = [_operation_key(op) for op in validated_ops]
        _LOGGER.info(
            f"Raw detected: {len(detected_ops)}  |  Validated: {len(validated_ops)} operations"
        )

        tp, fp, fn = compare_operations(validated_keys, gt_keys)
        scores = compute_scores(tp, fp, fn)

        r = AiotResult(
            aiot,
            len(gt_keys),
            len(detected_ops),
            len(validated_ops),
            tp,
            fp,
            fn,
            scores,
            elapsed_seconds=elapsed,
            usage=usage,
            cost_usd=cost,
        )
        results.append(r)

        total_tp, total_fp, total_fn = _log_and_accumulate(r, total_tp, total_fp, total_fn)
        _LOGGER.info(
            "  Time: %.1fs  |  Tokens: %d  |  Cost: $%.4f", elapsed, usage.total_tokens, cost
        )

    overall = compute_scores(total_tp, total_fp, total_fn)
    _LOGGER.info("\n%s", "=" * 50)
    _LOGGER.info("OVERALL (%s)", model_key)
    _LOGGER.info("%s", "=" * 50)
    _LOGGER.info("  TP=%d  FP=%d  FN=%d", total_tp, total_fp, total_fn)
    _LOGGER.info("  Precision: %.3f", overall.precision)
    _LOGGER.info("  Recall:    %.3f", overall.recall)
    _LOGGER.info("  F1:        %.3f", overall.f1)

    if args.save_score and eval_subdir is not None:
        xlsx_path = eval_subdir / "scores.xlsx"
        _write_xlsx(results, overall, total_tp, total_fp, total_fn, model_key, xlsx_path)
        _LOGGER.info("\nResults exported to %s", xlsx_path)

    return 0


if __name__ == "__main__":
    sys.exit(main())
